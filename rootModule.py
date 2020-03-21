#!/usr/bin/python3
# -*- coding:UTF-8 -*-
from yaml import load
from yaml import warnings
from yaml import dump
from openpyxl import load_workbook
from os import path
from string import ascii_uppercase,ascii_lowercase
from pyVim.connect import SmartConnect
from pyVmomi import vim
from binascii import a2b_hex
from time import strftime
import json, ssl


class vsphereSDK():
    def vsphereConfig(self, configStr):
        confStr = str(configStr)
        configFile = path.join(path.dirname(path.abspath(__file__)), confStr)
        warnings({'YAMLLoadWarning': False})
        config = load(open(configFile))
        return config

    def loginContent_var(self):
        '''Content Variable'''
        hosts_data = self.vsphereConfig(configStr='Conf/Account.yml')
        s = ssl.SSLContext(ssl.PROTOCOL_TLSv1_2)
        s.verify_mode = ssl.CERT_NONE
        for key in hosts_data.keys():
            si_instance = SmartConnect(host=hosts_data.get(key)['host'],
                                       user=hosts_data.get(key)['user'],
                                       pwd=bytes.decode(a2b_hex(hosts_data.get(key)['pass'])),
                                       sslContext=s
                                       )
            content = si_instance.RetrieveContent()
        return content

    def get_hostSystem_obj(self):
        content = self.loginContent_var()
        container = content.viewManager.CreateContainerView(content.rootFolder, [vim.HostSystem], True)
        obj = container.view
        container.Destroy()
        return obj

    def sheetTable_Style(self, ex_str):
        Ex_str = str(ex_str)
        tmpL0 = ascii_uppercase
        with open(Ex_str, 'rb') as Wf:
            Fo = load_workbook(Wf)
            ws = Fo.get_sheet_names()
            for w in ws:
                Wstr = w.encode('unicode-escape').decode('gbk')
                wb = Fo[Wstr]
                for s in tmpL0[:wb.max_column]:
                    wb.column_dimensions[str(s)].width = 50
        Fo.save(Ex_str)

    def jsonPrint(self, outData):
        if 'list' in str(type(outData)):
            print(json.dumps(outData, indent=len(outData)))
        elif 'dict' in str(type(outData)):
            print(json.dumps(outData, indent=len(outData.keys())))
        else:
            print('')

    def listTodict(self, dictStr, listStyle):
        valDict = dict()
        dictStr = str(dictStr)
        for itmVal in range(len(listStyle)):
            valDict["{};{}".format(dictStr, itmVal)] = listStyle[itmVal]
        return valDict

    def ymlStore(self, infoTag, dirVar):
        with open('tmp{}_{}.yml'.format(str(infoTag), strftime("%H%M")), 'a') as tf:
            dump(dir(dirVar), tf)

    def vmGuest_api(self):
        contentData = self.loginContent_var()
        for child_id in contentData.rootFolder.childEntity:
            childList = list(child_id.hostFolder.childEntity)
            host_list = [('vimHost:{}'.format(childItm),childList[childItm].host) for childItm in range(len(childList))]
        return host_list
    def yamlPath(self,setType,setName):
        if 'dict' in str(type(setType)):
            pathName = path.join(path.dirname(path.abspath(__file__)), str(setName))
            with open(str(pathName), 'w') as set_f:
                dump(setType, set_f)
        elif 'list' in str(type(setType)):
            pathName = path.join(path.dirname(path.abspath(__file__)), str(setName))
            with open(str(pathName), 'w') as str_f:
                dump(setType, str_f)
        else:
            print("武功再高也怕菜刀，衣服再叼一砖拍倒")



    def login_varStub(self):
        '''si_instance:_stub'''
        hosts_data = self.vsphereConfig(configStr='Conf/Account.yml')
        s = ssl.SSLContext(ssl.PROTOCOL_TLSv1_2)
        s.verify_mode = ssl.CERT_NONE
        for key in hosts_data.keys():
            si_instance = SmartConnect(host=hosts_data.get(key)['host'],
                                       user=hosts_data.get(key)['user'],
                                       pwd=bytes.decode(a2b_hex(hosts_data.get(key)['pass'])),
                                       sslContext=s
                                       )
            stub_var = si_instance._stub
        return stub_var

    def login_si(self):
        '''si_instance:_stub'''
        hosts_data = self.vsphereConfig(configStr='Conf/Account.yml')
        s = ssl.SSLContext(ssl.PROTOCOL_TLSv1_2)
        s.verify_mode = ssl.CERT_NONE
        for key in hosts_data.keys():
            si_instance = SmartConnect(host=hosts_data.get(key)['host'],
                                       user=hosts_data.get(key)['user'],
                                       pwd=bytes.decode(a2b_hex(hosts_data.get(key)['pass'])),
                                       sslContext=s
                                       )
        return si_instance

    def formatPatch(self,pathStyle):
        return path.join(path.dirname(path.abspath(__file__)), str(pathStyle))

    def dirFormat(self,style):
        caseSeq = list()
        for tag in ascii_lowercase:
            for case in dir(style):
                if str(case).startswith(tag) is True:
                    caseSeq.append(case)
        return caseSeq

def main():
    sdkVar=vsphereSDK()
    print(sdkVar.login_si())
    print(sdkVar.login_varStub())

    # sdkVar.jsonStyle_output(
    #     printData=sdkVar.vmGuest_api()
    # )


if __name__ == '__main__':
    main()
